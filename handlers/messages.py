import asyncio
import base64
import csv
import io
import logging
import os
import re

from telegram import Update
from telegram.ext import ContextTypes
from telegram.constants import ParseMode

from config import is_allowed, TTS_ENABLED
from agent import run_agent
from formatting import markdown_to_telegram_html, smart_split
import db

logger = logging.getLogger(__name__)


async def _keep_typing(bot, chat_id: int):
    """Send typing indicator every 4 seconds until cancelled."""
    try:
        while True:
            await bot.send_chat_action(chat_id=chat_id, action="typing")
            await asyncio.sleep(4)
    except asyncio.CancelledError:
        pass


async def _send_reply(update: Update, text: str, status_message=None):
    """Send formatted reply, falling back to plain text if HTML fails."""
    if not text:
        text = "(empty response)"

    # Delete status message if it exists
    if status_message:
        try:
            await status_message.delete()
        except Exception:
            pass

    html_text = markdown_to_telegram_html(text)
    chunks = smart_split(html_text)

    for chunk in chunks:
        try:
            await update.message.reply_text(
                chunk,
                parse_mode=ParseMode.HTML,
                disable_web_page_preview=True,
            )
        except Exception:
            # Fallback to plain text if HTML parsing fails
            plain_chunks = smart_split(text)
            for plain_chunk in plain_chunks:
                await update.message.reply_text(plain_chunk)
            break


async def _send_tts_audio(update: Update, reply: str):
    """Detect TTS_AUDIO_FILE sentinel in reply and send voice note."""
    tts_match = re.search(r'TTS_AUDIO_FILE:([^:]+):(\d+)', reply)
    if tts_match:
        filepath = tts_match.group(1)
        if os.path.exists(filepath):
            try:
                with open(filepath, "rb") as f:
                    await update.message.reply_voice(voice=f)
            except Exception as e:
                logger.error(f"Error sending TTS audio: {e}")
        return True
    return False


def _check_pending_confirmation(chat_id: int, user_message: str) -> str | None:
    """Check if user is responding to a pending confirmation."""
    pending = db.get_pending_confirmation(chat_id)
    if not pending:
        return None

    lower = user_message.lower().strip()
    affirmative = {"yes", "y", "confirm", "approve", "ok", "go ahead", "do it", "sure", "yep", "yeah"}
    negative = {"no", "n", "cancel", "deny", "stop", "don't", "nope", "nah", "abort"}

    if lower in affirmative:
        db.update_confirmation_status(pending["id"], "approved")
        return f"Confirmation #{pending['id']} approved. Proceeding with: {pending['action']} — {pending['details']}"
    elif lower in negative:
        db.update_confirmation_status(pending["id"], "denied")
        return f"Confirmation #{pending['id']} denied. Cancelled: {pending['action']}"

    return None


async def handle_custom_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle custom commands defined in the database (catch-all for unknown /commands)."""
    if not is_allowed(update.effective_user.id):
        await update.message.reply_text("Unauthorized.")
        return

    text = update.message.text or ""
    parts = text[1:].split(None, 1)
    cmd_name = parts[0].lower() if parts else ""
    cmd_input = parts[1] if len(parts) > 1 else ""

    cmd = db.get_custom_command_by_name(cmd_name)
    if not cmd:
        await update.message.reply_text(f"Unknown command: /{cmd_name}\nUse /help to see available commands.")
        return

    message = cmd["prompt_template"].replace("{input}", cmd_input).strip()

    chat_id = update.effective_chat.id
    typing_task = asyncio.create_task(_keep_typing(context.bot, chat_id))
    status_message = None

    async def status_callback(status_text: str):
        nonlocal status_message
        try:
            if status_message:
                await status_message.edit_text(f"_{status_text}_", parse_mode=ParseMode.MARKDOWN_V2)
            else:
                status_message = await update.message.reply_text(
                    f"_{status_text}_", parse_mode=ParseMode.MARKDOWN_V2
                )
        except Exception:
            pass

    try:
        reply = await run_agent(chat_id, message, status_callback=status_callback)
        # Check for TTS audio in response
        had_tts = await _send_tts_audio(update, reply)
        # Strip TTS sentinel from text reply
        clean_reply = re.sub(r'TTS_AUDIO_FILE:[^\s]+', '', reply).strip()
        if clean_reply:
            await _send_reply(update, clean_reply, status_message)
        elif status_message and not had_tts:
            await _send_reply(update, reply, status_message)
        elif status_message:
            try:
                await status_message.delete()
            except Exception:
                pass
        await _send_generated_files(chat_id, reply, update)
    except Exception as e:
        logger.error(f"Error in handle_custom_command: {e}", exc_info=True)
        await update.message.reply_text("Something went wrong. Try again.")
    finally:
        typing_task.cancel()


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        await update.message.reply_text("Unauthorized.")
        return

    chat_id = update.effective_chat.id
    user_message = update.message.text

    # Check for pending confirmation response
    confirmation_result = _check_pending_confirmation(chat_id, user_message)
    if confirmation_result:
        # Re-inject confirmation context and continue the agent
        user_message = confirmation_result

    typing_task = asyncio.create_task(_keep_typing(context.bot, chat_id))
    status_message = None

    async def status_callback(status_text: str):
        nonlocal status_message
        try:
            if status_message:
                await status_message.edit_text(f"_{status_text}_", parse_mode=ParseMode.MARKDOWN_V2)
            else:
                status_message = await update.message.reply_text(
                    f"_{status_text}_", parse_mode=ParseMode.MARKDOWN_V2
                )
        except Exception:
            pass

    try:
        reply = await run_agent(chat_id, user_message, status_callback=status_callback)
        # Check for TTS audio in response
        had_tts = await _send_tts_audio(update, reply)
        # Strip TTS sentinel from text reply
        clean_reply = re.sub(r'TTS_AUDIO_FILE:[^\s]+', '', reply).strip()
        if clean_reply:
            await _send_reply(update, clean_reply, status_message)
        elif status_message and not had_tts:
            await _send_reply(update, reply, status_message)
        elif status_message:
            try:
                await status_message.delete()
            except Exception:
                pass
        await _send_generated_files(chat_id, reply, update)
    except Exception as e:
        logger.error(f"Error in handle_text: {e}", exc_info=True)
        await update.message.reply_text("Something went wrong. Try again.")
    finally:
        typing_task.cancel()


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        await update.message.reply_text("Unauthorized.")
        return

    chat_id = update.effective_chat.id
    typing_task = asyncio.create_task(_keep_typing(context.bot, chat_id))

    try:
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        photo_bytes = await file.download_as_bytearray()
        b64 = base64.b64encode(photo_bytes).decode("utf-8")

        caption = update.message.caption or "What's in this image?"

        content = [
            {"type": "text", "text": caption},
            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
        ]

        reply = await run_agent(chat_id, content)
        await _send_reply(update, reply)
    except Exception as e:
        logger.error(f"Error in handle_photo: {e}", exc_info=True)
        await update.message.reply_text("Something went wrong processing that image.")
    finally:
        typing_task.cancel()


# ======================== FILE EXTRACTION ========================

# Extension → (label, language for code fence)
CODE_EXTENSIONS = {
    # Arduino / Embedded
    ".ino": ("Arduino Sketch", "cpp"), ".pde": ("Processing/Arduino", "cpp"),
    # C / C++
    ".c": ("C", "c"), ".h": ("C Header", "c"), ".cpp": ("C++", "cpp"),
    ".hpp": ("C++ Header", "cpp"), ".cc": ("C++", "cpp"), ".cxx": ("C++", "cpp"),
    # Python
    ".py": ("Python", "python"), ".pyw": ("Python", "python"), ".pyi": ("Python Stub", "python"),
    # JavaScript / TypeScript
    ".js": ("JavaScript", "javascript"), ".jsx": ("JSX", "jsx"),
    ".ts": ("TypeScript", "typescript"), ".tsx": ("TSX", "tsx"),
    ".mjs": ("JavaScript Module", "javascript"), ".cjs": ("CommonJS", "javascript"),
    # Web
    ".html": ("HTML", "html"), ".htm": ("HTML", "html"), ".css": ("CSS", "css"),
    ".scss": ("SCSS", "scss"), ".sass": ("Sass", "sass"), ".less": ("Less", "less"),
    ".vue": ("Vue", "vue"), ".svelte": ("Svelte", "svelte"),
    # Data / Config
    ".json": ("JSON", "json"), ".yaml": ("YAML", "yaml"), ".yml": ("YAML", "yaml"),
    ".toml": ("TOML", "toml"), ".xml": ("XML", "xml"), ".ini": ("INI", "ini"),
    ".cfg": ("Config", "ini"), ".conf": ("Config", ""), ".env": ("Environment", ""),
    ".properties": ("Properties", "properties"),
    # Shell / Scripts
    ".sh": ("Shell Script", "bash"), ".bash": ("Bash Script", "bash"),
    ".zsh": ("Zsh Script", "bash"), ".fish": ("Fish Script", "fish"),
    ".bat": ("Batch File", "batch"), ".cmd": ("Command Script", "batch"),
    ".ps1": ("PowerShell", "powershell"),
    # Systems / Low-level
    ".rs": ("Rust", "rust"), ".go": ("Go", "go"), ".java": ("Java", "java"),
    ".kt": ("Kotlin", "kotlin"), ".kts": ("Kotlin Script", "kotlin"),
    ".swift": ("Swift", "swift"), ".m": ("Objective-C", "objectivec"),
    ".scala": ("Scala", "scala"), ".zig": ("Zig", "zig"),
    # Ruby / PHP / Perl
    ".rb": ("Ruby", "ruby"), ".php": ("PHP", "php"), ".pl": ("Perl", "perl"),
    ".pm": ("Perl Module", "perl"),
    # Databases
    ".sql": ("SQL", "sql"), ".prisma": ("Prisma", "prisma"),
    # Docs / Markup
    ".md": ("Markdown", "markdown"), ".rst": ("reStructuredText", "rst"),
    ".tex": ("LaTeX", "latex"), ".adoc": ("AsciiDoc", "asciidoc"),
    # DevOps / CI
    ".dockerfile": ("Dockerfile", "dockerfile"),
    ".tf": ("Terraform", "hcl"), ".hcl": ("HCL", "hcl"),
    ".nix": ("Nix", "nix"),
    # Data Science
    ".r": ("R", "r"), ".jl": ("Julia", "julia"), ".m": ("MATLAB", "matlab"),
    # Build
    ".cmake": ("CMake", "cmake"), ".gradle": ("Gradle", "groovy"),
    ".sbt": ("SBT", "scala"),
    # Misc
    ".proto": ("Protocol Buffers", "protobuf"), ".graphql": ("GraphQL", "graphql"),
    ".gql": ("GraphQL", "graphql"), ".wasm": ("WebAssembly", ""),
    ".lua": ("Lua", "lua"), ".dart": ("Dart", "dart"),
    ".ex": ("Elixir", "elixir"), ".exs": ("Elixir Script", "elixir"),
    ".erl": ("Erlang", "erlang"), ".hs": ("Haskell", "haskell"),
    ".clj": ("Clojure", "clojure"), ".lisp": ("Lisp", "lisp"),
    ".v": ("Verilog", "verilog"), ".vhd": ("VHDL", "vhdl"), ".vhdl": ("VHDL", "vhdl"),
    ".sv": ("SystemVerilog", "systemverilog"),
    ".asm": ("Assembly", "asm"), ".s": ("Assembly", "asm"),
    ".S": ("Assembly", "asm"),
    # IoT / Embedded config
    ".platformio.ini": ("PlatformIO Config", "ini"),
}

# Files detected by exact name
NAMED_FILES = {
    "Makefile": ("Makefile", "makefile"), "makefile": ("Makefile", "makefile"),
    "CMakeLists.txt": ("CMake", "cmake"),
    "Dockerfile": ("Dockerfile", "dockerfile"),
    "docker-compose.yml": ("Docker Compose", "yaml"),
    "docker-compose.yaml": ("Docker Compose", "yaml"),
    ".gitignore": ("Git Ignore", ""), ".gitmodules": ("Git Modules", ""),
    ".editorconfig": ("Editor Config", "ini"),
    "Vagrantfile": ("Vagrantfile", "ruby"),
    "Rakefile": ("Rakefile", "ruby"),
    "Gemfile": ("Gemfile", "ruby"),
    "Procfile": ("Procfile", ""),
    "platformio.ini": ("PlatformIO Config", "ini"),
    "requirements.txt": ("Python Requirements", ""),
    "package.json": ("NPM Package", "json"),
    "tsconfig.json": ("TypeScript Config", "json"),
    "Cargo.toml": ("Cargo Config", "toml"),
    "go.mod": ("Go Module", ""),
    "go.sum": ("Go Sum", ""),
}

# Binary file extensions we can describe but not read as text
BINARY_EXTENSIONS = {
    ".bin", ".hex", ".elf", ".o", ".obj", ".a", ".lib", ".so", ".dll", ".dylib",
    ".exe", ".out", ".class", ".jar", ".war",
    ".stl", ".obj", ".step", ".stp", ".iges", ".igs",  # 3D / CAD
    ".pcb", ".sch", ".brd", ".kicad_pcb", ".kicad_sch",  # PCB / EDA (some are text)
    ".gerber", ".gbr", ".drl",  # Gerber
    ".mp3", ".wav", ".ogg", ".flac", ".aac",  # Audio
    ".mp4", ".avi", ".mov", ".mkv", ".webm",  # Video
    ".ttf", ".otf", ".woff", ".woff2",  # Fonts
    ".sqlite", ".db",  # Databases
}

MAX_TEXT_LEN = 15000
MAX_TEXT_FALLBACK_LEN = 10000


def _detect_file_type(filename: str) -> tuple[str, str] | None:
    """Detect file type from filename. Returns (label, language) or None."""
    basename = os.path.basename(filename)
    if basename in NAMED_FILES:
        return NAMED_FILES[basename]
    ext = os.path.splitext(filename.lower())[1]
    if ext in CODE_EXTENSIONS:
        return CODE_EXTENSIONS[ext]
    return None


def _decode_text(file_bytes: bytes) -> str | None:
    """Try to decode bytes as text with multiple encodings."""
    for encoding in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            return file_bytes.decode(encoding)
        except (UnicodeDecodeError, ValueError):
            continue
    return None


def _extract_text_file(file_bytes: bytes, filename: str) -> tuple[str, str]:
    """Extract text from a code/text file. Returns (label, formatted_text)."""
    detected = _detect_file_type(filename)
    text = _decode_text(file_bytes)
    if text is None:
        return "Binary File", f"(Binary file, {len(file_bytes)} bytes — cannot display as text)"

    if len(text) > MAX_TEXT_LEN:
        text = text[:MAX_TEXT_LEN] + "\n...(truncated)"

    if detected:
        label, lang = detected
        if lang:
            return label, f"```{lang}\n{text}\n```"
        return label, text
    return "File", text


def _extract_pdf_text(file_bytes: bytes) -> str:
    """Extract text from a PDF file using PyPDF2."""
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(file_bytes))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"--- Page {i + 1} ---\n{text}")
    return "\n\n".join(pages) if pages else "(No extractable text in PDF)"


def _extract_csv_text(file_bytes: bytes) -> str:
    """Read CSV and format as a markdown table."""
    text = _decode_text(file_bytes) or file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return "(Empty CSV)"

    header = rows[0]
    lines = ["| " + " | ".join(header) + " |"]
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in rows[1:101]:
        lines.append("| " + " | ".join(row) + " |")
    result = "\n".join(lines)
    if len(rows) > 101:
        result += f"\n\n(Showing 100 of {len(rows) - 1} rows)"
    return result


def _extract_docx_text(file_bytes: bytes) -> str:
    """Extract text from a Word .docx file."""
    from docx import Document
    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs) if paragraphs else "(No extractable text in document)"


def _extract_excel_text(file_bytes: bytes) -> str:
    """Read Excel (.xlsx/.xls) and format as a markdown table."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        if not rows:
            continue

        header = rows[0]
        lines = [f"**Sheet: {sheet_name}**", "| " + " | ".join(header) + " |"]
        lines.append("| " + " | ".join("---" for _ in header) + " |")
        for row in rows[1:101]:
            lines.append("| " + " | ".join(row) + " |")
        if len(rows) > 101:
            lines.append(f"\n(Showing 100 of {len(rows) - 1} rows)")
        sheets.append("\n".join(lines))
    wb.close()
    return "\n\n".join(sheets) if sheets else "(Empty spreadsheet)"


def _extract_zip_contents(file_bytes: bytes) -> str:
    """List contents of a ZIP archive and extract text from small text files."""
    import zipfile
    zf = zipfile.ZipFile(io.BytesIO(file_bytes))
    entries = zf.namelist()
    lines = [f"**Archive contents ({len(entries)} files):**"]
    for name in entries[:50]:
        info = zf.getinfo(name)
        size = info.file_size
        lines.append(f"- `{name}` ({size:,} bytes)")
    if len(entries) > 50:
        lines.append(f"- ... and {len(entries) - 50} more files")

    # Try to extract small text files
    text_extensions = {".ino", ".c", ".h", ".cpp", ".py", ".js", ".ts", ".json", ".yaml", ".yml",
                       ".toml", ".txt", ".md", ".html", ".css", ".xml", ".sh", ".cfg", ".ini", ".sql"}
    extracted = []
    total_extracted = 0
    for name in entries:
        ext = os.path.splitext(name.lower())[1]
        info = zf.getinfo(name)
        if ext in text_extensions and info.file_size < 50000 and info.file_size > 0 and total_extracted < 30000:
            try:
                content = zf.read(name).decode("utf-8", errors="replace")
                detected = _detect_file_type(name)
                lang = detected[1] if detected else ""
                fence = f"```{lang}\n{content}\n```" if lang else content
                extracted.append(f"\n**{name}:**\n{fence}")
                total_extracted += len(content)
            except Exception:
                pass

    zf.close()
    result = "\n".join(lines)
    if extracted:
        result += "\n\n**Extracted files:**" + "\n".join(extracted)
    return result


def _extract_ipynb_text(file_bytes: bytes) -> str:
    """Extract content from a Jupyter notebook."""
    import json as _json
    nb = _json.loads(file_bytes.decode("utf-8"))
    cells = nb.get("cells", [])
    parts = []
    for i, cell in enumerate(cells):
        cell_type = cell.get("cell_type", "code")
        source = "".join(cell.get("source", []))
        if not source.strip():
            continue
        if cell_type == "markdown":
            parts.append(source)
        else:
            parts.append(f"```python\n{source}\n```")
        # Include outputs for code cells
        outputs = cell.get("outputs", [])
        for out in outputs:
            if "text" in out:
                parts.append("Output:\n" + "".join(out["text"]))
            elif "data" in out and "text/plain" in out["data"]:
                parts.append("Output:\n" + "".join(out["data"]["text/plain"]))
    return "\n\n".join(parts) if parts else "(Empty notebook)"


async def _send_generated_files(chat_id: int, reply: str, update: Update):
    """Check if the agent generated any files and send them via Telegram."""
    # Look for file IDs in the agent's tool output (pattern: "id=<number>")
    file_ids = re.findall(r'\(id=(\d+)\)', reply)
    for fid in file_ids:
        try:
            file_record = db.get_file_upload(int(fid))
            if file_record and os.path.exists(file_record["path"]):
                with open(file_record["path"], "rb") as f:
                    await update.message.reply_document(
                        document=f,
                        filename=file_record["filename"],
                    )
        except Exception as e:
            logger.error(f"Error sending generated file {fid}: {e}")


def _process_file(file_bytes: bytes, filename: str, mime: str) -> tuple[str | list, str]:
    """Process any file and return (content_for_agent, label).
    content_for_agent is either a string or a list (for vision/images)."""
    fn_lower = filename.lower()
    ext = os.path.splitext(fn_lower)[1]

    # Images → vision API
    if mime.startswith("image/"):
        b64 = base64.b64encode(file_bytes).decode("utf-8")
        return [
            {"type": "text", "text": "What's in this image?"},
            {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
        ], "Image"

    # PDF
    if mime == "application/pdf" or fn_lower.endswith(".pdf"):
        return _extract_pdf_text(file_bytes), "PDF"

    # CSV
    if mime == "text/csv" or fn_lower.endswith(".csv") or fn_lower.endswith(".tsv"):
        return _extract_csv_text(file_bytes), "CSV"

    # Excel
    if fn_lower.endswith((".xlsx", ".xlsm")) or "spreadsheetml" in mime:
        return _extract_excel_text(file_bytes), "Excel"

    # Word
    if fn_lower.endswith(".docx") or "wordprocessingml" in mime:
        return _extract_docx_text(file_bytes), "Word Document"

    # ZIP / archives
    if fn_lower.endswith((".zip", ".ino.zip")) or mime == "application/zip":
        return _extract_zip_contents(file_bytes), "Archive"

    # Jupyter notebook
    if fn_lower.endswith(".ipynb"):
        return _extract_ipynb_text(file_bytes), "Jupyter Notebook"

    # Known binary files — describe but don't try to read
    if ext in BINARY_EXTENSIONS:
        return f"(Binary file: {filename}, {len(file_bytes):,} bytes, type: {ext})", "Binary File"

    # Everything else — try as text (code files, configs, scripts, etc.)
    label, text = _extract_text_file(file_bytes, filename)
    return text, label


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        await update.message.reply_text("Unauthorized.")
        return

    chat_id = update.effective_chat.id
    typing_task = asyncio.create_task(_keep_typing(context.bot, chat_id))

    try:
        doc = update.message.document
        mime = doc.mime_type or ""
        filename = doc.file_name or "unknown"

        file = await context.bot.get_file(doc.file_id)
        file_bytes = bytes(await file.download_as_bytearray())
        caption = update.message.caption or ""

        content, label = _process_file(file_bytes, filename, mime)

        if isinstance(content, list):
            # Vision content (images)
            if caption:
                content[0]["text"] = caption
            reply = await run_agent(chat_id, content)
        else:
            if len(content) > MAX_TEXT_LEN:
                content = content[:MAX_TEXT_LEN] + "\n...(truncated)"
            user_message = f"[{label}: {filename}]\n{content}"
            if caption:
                user_message = f"{caption}\n\n{user_message}"
            reply = await run_agent(chat_id, user_message)

        await _send_reply(update, reply)
    except Exception as e:
        logger.error(f"Error in handle_document: {e}", exc_info=True)
        await update.message.reply_text("Something went wrong processing that file.")
    finally:
        typing_task.cancel()


async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        await update.message.reply_text("Unauthorized.")
        return

    chat_id = update.effective_chat.id
    typing_task = asyncio.create_task(_keep_typing(context.bot, chat_id))

    try:
        voice = update.message.voice or update.message.audio
        file = await context.bot.get_file(voice.file_id)
        file_bytes = await file.download_as_bytearray()

        # Write to temp file for Whisper API
        import tempfile, os
        with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name

        try:
            from openai import AsyncOpenAI
            from config import OPENAI_API_KEY
            client = AsyncOpenAI(api_key=OPENAI_API_KEY)

            with open(tmp_path, "rb") as audio_file:
                transcription = await client.audio.transcriptions.create(
                    model="whisper-1",
                    file=audio_file,
                )
            transcript = transcription.text
        finally:
            os.unlink(tmp_path)

        if not transcript:
            await update.message.reply_text("Couldn't transcribe that audio.")
            return

        # If TTS is enabled, hint the agent to respond with voice
        if TTS_ENABLED:
            transcript = f"{transcript}\n\n[User sent a voice message. Consider using text_to_speech to respond with a voice note.]"

        status_message = None

        async def status_callback(status_text: str):
            nonlocal status_message
            try:
                if status_message:
                    await status_message.edit_text(f"_{status_text}_", parse_mode=ParseMode.MARKDOWN_V2)
                else:
                    status_message = await update.message.reply_text(
                        f"_{status_text}_", parse_mode=ParseMode.MARKDOWN_V2
                    )
            except Exception:
                pass

        reply = await run_agent(chat_id, transcript, status_callback=status_callback)
        # Check for TTS audio in response
        had_tts = await _send_tts_audio(update, reply)
        # Strip TTS sentinel from text reply
        clean_reply = re.sub(r'TTS_AUDIO_FILE:[^\s]+', '', reply).strip()
        if clean_reply:
            await _send_reply(update, clean_reply, status_message)
        elif status_message and not had_tts:
            await _send_reply(update, reply, status_message)
        elif status_message:
            try:
                await status_message.delete()
            except Exception:
                pass
    except Exception as e:
        logger.error(f"Error in handle_voice: {e}", exc_info=True)
        await update.message.reply_text("Something went wrong processing that voice message.")
    finally:
        typing_task.cancel()
